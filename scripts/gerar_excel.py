#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera a PLANILHA DE ESTUDO viva (.xlsx) — versão com estética da marca."""
import argparse, json, os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.chart import BarChart, Reference

_LOGO_DEFAULT = Path(__file__).resolve().parent.parent / "assets" / "logo.png"
LOGO = os.environ.get("EDITAL_LOGO") or str(_LOGO_DEFAULT)  # opcional: marca própria do usuário

AZUL = "2563EB"; AZUL_DARK = "1D4ED8"; AZUL_LIGHT = "60A5FA"; SLATE = "1E293B"
ZEBRA = "F5F8FD"; BRANCO = "FFFFFF"
F_VERDE = "C6EFCE"; F_AZUL = "DCE9FB"; F_AMARELO = "FFF1C2"; F_VERM = "FBE0E0"
F_INC_ALTA = "F9C9C9"; F_INC_MED = "FDE7B5"; F_INC_BAIXA = "E7EDF5"

STATUS = ["A estudar", "Estudando", "Estudado", "Revisão 1", "Revisão 2", "Revisão 3"]
INCID = ["Sem dado", "Alta", "Média", "Baixa"]
PRIOR = ["Alta", "Média", "Baixa"]
HEADERS = ["Disciplina", "Nº", "Nível", "Conteúdo", "Status", "Incidência", "Prioridade", "Anotações"]

THIN = Side(style="thin", color="DCE3EC")
SEP = Side(style="medium", color=AZUL_LIGHT)
HEAD_ROW = 3


def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def _banner(ws, data):
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 20
    ws.merge_cells("B1:H1"); ws.merge_cells("B2:H2")
    for col in range(2, 9):
        ws.cell(1, col).fill = _fill(AZUL)
        ws.cell(2, col).fill = _fill(AZUL_DARK)
    ws["B1"] = "Edital Verticalizado — plano de estudo"
    ws["B1"].font = Font(bold=True, color=BRANCO, size=15)
    ws["B1"].alignment = Alignment(vertical="center", indent=1)
    sub = " · ".join(x for x in [data.get("concurso"), data.get("cargo")] if x)
    ws["B2"] = sub or "Plano de estudo"
    ws["B2"].font = Font(color="DBE7FF", size=10.5)
    ws["B2"].alignment = Alignment(vertical="center", indent=1)
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(LOGO); img.height = 44; img.width = 44
        ws.add_image(img, "A1")
    except Exception:
        pass  # sem logo: o título do banner se sustenta sozinho


def _header(ws):
    for c, name in enumerate(HEADERS, start=1):
        cell = ws.cell(HEAD_ROW, c, name)
        cell.fill = _fill(AZUL); cell.font = Font(bold=True, color=BRANCO, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.cell(HEAD_ROW, 9, "_feito")


def build(data, out):
    wb = Workbook(); ws = wb.active; ws.title = "Verticalizado"
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 110
    ws.sheet_properties.tabColor = AZUL
    _banner(ws, data); _header(ws)

    r = HEAD_ROW + 1
    for di, disc in enumerate(data.get("disciplinas", [])):
        nome = disc.get("nome", "")
        block = _fill(ZEBRA if di % 2 == 0 else BRANCO)
        first = True
        for it in disc.get("itens", []):
            nivel = int(it.get("nivel", 1))
            vals = [nome, it.get("numero", ""), nivel,
                    ("   " * (nivel - 1)) + it.get("texto", ""), "A estudar", "Sem dado", "", ""]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v)
                cell.fill = block
                top = SEP if first else THIN
                cell.border = Border(left=THIN, right=THIN, top=top, bottom=THIN)
                if c in (2, 3, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=(c in (4, 8)))
            # assunto (nível 1) em negrito
            if nivel == 1:
                ws.cell(r, 1).font = Font(bold=True, color=SLATE, size=10)
                ws.cell(r, 4).font = Font(bold=True, color=SLATE)
            else:
                ws.cell(r, 1).font = Font(color="64748B", size=9)
                ws.cell(r, 4).font = Font(color="334155")
            ws.cell(r, 9, f'=IF(OR($E{r}="Estudado",LEFT($E{r},7)="Revisão"),1,0)').fill = block
            first = False
            r += 1
    last = r - 1

    for col, w in {"A": 30, "B": 7, "C": 7, "D": 74, "E": 13, "F": 12, "G": 11, "H": 26, "I": 6}.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["I"].hidden = True
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{HEAD_ROW}:H{last}"

    def dv(col, opts):
        d = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=True)
        d.add(f"{col}{HEAD_ROW+1}:{col}{last}"); ws.add_data_validation(d)
    dv("E", STATUS); dv("F", INCID); dv("G", PRIOR)

    e = f"E{HEAD_ROW+1}:E{last}"; f = f"F{HEAD_ROW+1}:F{last}"
    rules = [
        (e, f'$E{HEAD_ROW+1}="Estudado"', F_VERDE),
        (e, f'LEFT($E{HEAD_ROW+1},7)="Revisão"', F_AZUL),
        (e, f'$E{HEAD_ROW+1}="Estudando"', F_AMARELO),
        (e, f'$E{HEAD_ROW+1}="A estudar"', F_VERM),
        (f, f'$F{HEAD_ROW+1}="Alta"', F_INC_ALTA),
        (f, f'$F{HEAD_ROW+1}="Média"', F_INC_MED),
        (f, f'$F{HEAD_ROW+1}="Baixa"', F_INC_BAIXA),
    ]
    for rng, formula, color in rules:
        ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=_fill(color)))

    _resumo(wb, data, last); _ajuda(wb)
    wb.save(out)
    return last


def _resumo(wb, data, last):
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = AZUL_LIGHT
    ws["A1"] = "Resumo de cobertura"; ws["A1"].font = Font(bold=True, size=15, color=SLATE)
    ws["A2"] = " · ".join(x for x in [data.get("concurso"), data.get("cargo")] if x)
    ws["A2"].font = Font(italic=True, color="64748B")
    hdr_row = 4
    for c, name in enumerate(["Disciplina", "Itens", "Concluídos", "% Cobertura"], start=1):
        cell = ws.cell(hdr_row, c, name); cell.fill = _fill(AZUL)
        cell.font = Font(bold=True, color=BRANCO); cell.alignment = Alignment(horizontal="center")
    row = hdr_row + 1
    first_data = row
    for disc in data.get("disciplinas", []):
        ws.cell(row, 1, disc.get("nome", ""))
        ws.cell(row, 2, f"=COUNTIF(Verticalizado!$A${HEAD_ROW+1}:$A${last},A{row})")
        ws.cell(row, 3, f"=SUMIFS(Verticalizado!$I${HEAD_ROW+1}:$I${last},Verticalizado!$A${HEAD_ROW+1}:$A${last},A{row})")
        p = ws.cell(row, 4, f"=IF(B{row}=0,0,C{row}/B{row})"); p.number_format = "0%"
        for c in range(1, 5):
            ws.cell(row, c).border = Border(bottom=THIN)
        row += 1
    last_data = row - 1
    ws.cell(row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(row, 2, f"=SUM(B{first_data}:B{last_data})").font = Font(bold=True)
    ws.cell(row, 3, f"=SUM(C{first_data}:C{last_data})").font = Font(bold=True)
    tp = ws.cell(row, 4, f"=IF(B{row}=0,0,C{row}/B{row})"); tp.number_format = "0%"; tp.font = Font(bold=True)
    ws.conditional_formatting.add(f"D{first_data}:D{row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=AZUL_LIGHT))
    for col, w in {"A": 34, "B": 9, "C": 11, "D": 13}.items():
        ws.column_dimensions[col].width = w

    chart = BarChart(); chart.type = "bar"; chart.title = "% de cobertura por disciplina"
    chart.height = max(7, 0.5 * (last_data - first_data + 1)); chart.width = 18
    chart.legend = None
    vals = Reference(ws, min_col=4, min_row=hdr_row, max_row=last_data)
    cats = Reference(ws, min_col=1, min_row=first_data, max_row=last_data)
    chart.add_data(vals, titles_from_data=True); chart.set_categories(cats)
    chart.y_axis.numFmt = "0%"; chart.x_axis.delete = False; chart.y_axis.delete = False
    ws.add_chart(chart, "F4")


def _ajuda(wb):
    ws = wb.create_sheet("Como usar")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = "94A3B8"
    linhas = [
        ("Como usar esta planilha", True),
        ("", False),
        ("1. Aba 'Verticalizado': cada linha é um item do edital, na ordem original.", False),
        ("2. 'Status': marque seu progresso (A estudar → Estudado → Revisão 1/2/3).", False),
        ("   As cores e a aba 'Resumo' (% de cobertura) atualizam sozinhas.", False),
        ("3. 'Incidência': começa como 'Sem dado' DE PROPÓSITO.", False),
        ("   A IA não chuta o que mais cai. Para preencher de verdade, junte os assuntos", False),
        ("   das últimas provas da SUA banca e classifique cada item em Alta/Média/Baixa.", False),
        ("4. 'Prioridade': ataque primeiro o que é Incidência Alta e ainda não estudou.", False),
        ("5. 'Anotações': links, dúvidas, mnemônicos.", False),
        ("", False),
        ("Princípio: a IA organiza, você decide. Ela não estuda no seu lugar — libera seu tempo.", False),
    ]
    for i, (txt, bold) in enumerate(linhas, start=1):
        c = ws.cell(i, 1, txt)
        c.font = Font(bold=bold, size=15 if bold else 11, color=SLATE if bold else "334155")
    ws.column_dimensions["A"].width = 100


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("json"); ap.add_argument("--out", "-o", required=True)
    args = ap.parse_args()
    data = json.loads(Path(args.json).read_text(encoding="utf-8"))
    last = build(data, args.out)
    print(f"OK -> {args.out}")
    print(f"Linhas de item: {last - HEAD_ROW} | Disciplinas: {len(data.get('disciplinas', []))} | Abas: Verticalizado, Resumo, Como usar")


if __name__ == "__main__":
    main()
